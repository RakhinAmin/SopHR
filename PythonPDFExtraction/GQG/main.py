# Import the PyMuPDF library (fitz) for handling PDF files
import fitz  
# Import the regular expression library (re) for pattern matching
import re  
# Import the os library for interacting with the operating system (e.g., checking file existence)
import os  
# Import the load_workbook function from openpyxl to load existing Excel files
from openpyxl import load_workbook  
# Import the Workbook class from openpyxl to create new Excel files
from openpyxl import Workbook  

# Define a class named PdfProcessor to encapsulate PDF processing logic
class PdfProcessor:
    # Class constants for section markers and regular expression patterns
    SECTION_START = "Fund\nUnits/Shares"  
    SECTION_END = "Total"  
    # Regular expression patterns for units/shares, price, and value
    UNITS_PATTERN = r"\d{1,3}(?:,\d{3})*(?:\.\d+)?"
    PRICE_PATTERN = r"(?<=\n)\d{1,2}\.\d{2}(?=\n)"  # Ensure it captures price on its own line
    VALUE_PATTERN = r"USD \d{1,3}(?:,\d{3})*\.\d{2}"

    # Constructor method to initialize the PdfProcessor instance with a PDF path
    def __init__(self, pdf_path):
        # Store the PDF path as an instance attribute
        self.pdf_path = pdf_path  
        # Extract text from the PDF and store it as an instance attribute
        self.text = self._extract_text()  

    # Private method to extract text from the PDF file
    def _extract_text(self):
        try:
            # Attempt to open the PDF file using PyMuPDF
            document = fitz.open(self.pdf_path)              
            # Extract text from each page and concatenate it into a single string
            return "".join(page.get_text() for page in document)         
        except Exception as e:
            # Raise a custom exception if PDF processing fails
            raise RuntimeError(f"PDF processing failed: {e}")  

    # Public method to extract financial data from the PDF text
    def extract_financial_data(self):
        # Find the relevant section in the text
        section = self._find_section()        
        # If the section is found, parse it for financial data
        return self._parse_section(section) if section else (None, None)  

    # Private helper method to find the section of interest in the text
    def _find_section(self):
        # Find the indices of the start and end markers in the text
        start = self.text.find(self.SECTION_START)  
        end = self.text.find(self.SECTION_END, start)          
        # If both markers are found, extract the section text
        return self.text[start:end] if start != -1 and end != -1 else None  

    # Private method to parse the section text for financial data
    def _parse_section(self, section):
        # Search for units/shares, price, and value patterns in the section text
        units = re.search(self.UNITS_PATTERN, section)  
        price = re.search(self.PRICE_PATTERN, section)  
        value = re.search(self.VALUE_PATTERN, section)  

        # Check if all patterns were matched
        if all((units, price, value)):
            # Extract and return the matched values along with headers
            return (
                ["Units/Shares", "Price", "Value in Share Class Currency"],
                [g.group(0).strip() for g in (units, price, value)]
            )
        
        # If any pattern is not matched, return None for both headers and values
        return None, None  


# Define a class named ExcelHandler to encapsulate Excel operations
class ExcelHandler:
    # Class constants for the default Excel filename and sheet name
    DEFAULT_FILENAME = 'GQG_Cleaned_Data.xlsx'  
    SHEET_NAME = 'GQG'  

    # Constructor method to initialize the ExcelHandler instance
    def __init__(self):
        # Initialize the workbook by loading or creating it
        self.wb = self._init_workbook()         
        # Get the worksheet with the specified name
        self.ws = self.wb[self.SHEET_NAME]  

    # Private method to initialize the workbook
    def _init_workbook(self):
        # Check if the Excel file already exists
        if os.path.exists(self.DEFAULT_FILENAME):
            # If it exists, load the workbook
            return load_workbook(self.DEFAULT_FILENAME)  
                
        # If it doesn't exist, create a new workbook
        wb = Workbook()         
        # Rename the default sheet to match the specified sheet name
        wb.active.title = self.SHEET_NAME         
        # Return the newly created workbook
        return wb  

    # Public method to append data to the Excel worksheet
    def append_data(self, headers, values):
        # Check if the worksheet is empty (i.e., new file)
        if self._is_new_file():
            # If it's empty, append the headers to the first row
            self.ws.append(headers)  
        
        # Append the values to the next available row
        self.ws.append(values)         
        # Save the changes to the workbook
        self.wb.save(self.DEFAULT_FILENAME)  

    # Private helper method to check if the worksheet is empty
    def _is_new_file(self):
        # Check if the worksheet has only one cell with no value
        return self.ws.max_row == 1 and self.ws.max_column == 1 and not self.ws['A1'].value  


# Client code usage
if __name__ == "__main__":
    # Specify the path to the PDF file to process
    pdf_path = r'C:\Users\Rakhin.Amin\Documents\PythonPDFExtraction\GQG\GQG - Capital Statement (Feb 24).pdf'  
    
    try:
        # Create a PdfProcessor instance with the specified PDF path
        processor = PdfProcessor(pdf_path)         
        # Extract financial data from the PDF
        headers, values = processor.extract_financial_data()  
        
        # Check if data was successfully extracted
        if headers and values:
            # Create an ExcelHandler instance
            excel_handler = ExcelHandler()            
            # Append the extracted data to the Excel file
            excel_handler.append_data(headers, values)            
            # Print a success message
            print("Data processed successfully")  
            
    except Exception as e:
        # Catch any exceptions and print the error message
        print(f"Processing error: {e}")  
